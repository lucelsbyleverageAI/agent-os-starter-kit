"""Excel processor utility for extracting calculated values from Excel files."""

import logging
import tempfile
from pathlib import Path
from typing import List, Optional
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

from .logging import get_logger

logger = get_logger(__name__)


def is_excel_url(url: str) -> bool:
    """Check if a URL points to an Excel file.

    Args:
        url: URL to check

    Returns:
        True if URL appears to be an Excel file
    """
    excel_extensions = ('.xlsx', '.xls', '.xlsm', '.xlsb')
    parsed = urlparse(url)
    path = parsed.path.lower()
    return any(path.endswith(ext) for ext in excel_extensions)


async def process_excel_url(url: str) -> str:
    """Download and process an Excel file from a URL.

    Downloads the Excel file, extracts calculated values using openpyxl,
    and returns the content as markdown formatted text.

    Args:
        url: URL of the Excel file to process

    Returns:
        Markdown formatted content from the Excel file

    Raises:
        Exception: If download or processing fails
    """
    temp_file = None
    try:
        # Download the file to a temporary location
        async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
            response = await client.get(url)
            response.raise_for_status()

            # Create temporary file with appropriate extension
            suffix = '.xlsx'
            for ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                if url.lower().endswith(ext):
                    suffix = ext
                    break

            temp_file = tempfile.NamedTemporaryFile(
                mode='wb',
                suffix=suffix,
                delete=False
            )
            temp_file.write(response.content)
            temp_file.close()

            # Process the Excel file
            content = _process_excel_file(temp_file.name)
            return content

    except Exception as e:
        logger.error(f"Error processing Excel file from {url}: {e}", exc_info=True)
        raise

    finally:
        # Clean up temporary file
        if temp_file:
            try:
                Path(temp_file.name).unlink(missing_ok=True)
            except Exception as e:
                logger.warning(f"Failed to delete temporary file: {e}")


def _process_excel_file(file_path: str) -> str:
    """Process an Excel file and extract content with calculated values.

    Args:
        file_path: Path to the Excel file

    Returns:
        Markdown formatted content
    """
    try:
        # Load workbook with data_only=True to get calculated values
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)

        # Process all sheets
        all_sheets_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_content = _process_sheet(sheet, sheet_name)

            if sheet_content.strip():  # Only include non-empty sheets
                all_sheets_content.append(sheet_content)

        # Combine all sheets
        if all_sheets_content:
            content = "\n\n---\n\n".join(all_sheets_content)
        else:
            content = "Empty Excel file"

        workbook.close()

        return content

    except Exception as e:
        logger.error(f"Error processing Excel file {file_path}: {e}", exc_info=True)
        raise


def _process_sheet(sheet, sheet_name: str) -> str:
    """Process a single Excel sheet and convert to markdown.

    Args:
        sheet: openpyxl worksheet object
        sheet_name: Name of the sheet

    Returns:
        Markdown representation of the sheet
    """
    lines = []

    # Add sheet header
    lines.append(f"# Sheet: {sheet_name}")
    lines.append("")

    # Find the actual data range (excluding empty rows/columns)
    data_rows = []
    max_col = 0

    for row in sheet.iter_rows(values_only=True):
        # Convert row values, handling None
        row_values = [_format_cell_value(val) for val in row]

        # Check if row has any non-empty values
        if any(val for val in row_values):
            data_rows.append(row_values)
            # Track maximum column with data
            for i, val in enumerate(row_values):
                if val:
                    max_col = max(max_col, i + 1)

    if not data_rows:
        lines.append("_Empty sheet_")
        return "\n".join(lines)

    # Trim rows to max_col
    trimmed_rows = [row[:max_col] for row in data_rows]

    # Convert to markdown table
    if trimmed_rows:
        # Create table
        table_md = _create_markdown_table(trimmed_rows)
        lines.append(table_md)

    return "\n".join(lines)


def _format_cell_value(value) -> str:
    """Format a cell value for display.

    Args:
        value: Cell value (can be None, number, string, etc.)

    Returns:
        Formatted string representation
    """
    if value is None:
        return ""
    elif isinstance(value, (int, float)):
        # Format numbers nicely
        if isinstance(value, float):
            # Check if it's a whole number
            if value.is_integer():
                return str(int(value))
            else:
                # Format with appropriate precision
                return f"{value:.2f}".rstrip('0').rstrip('.')
        return str(value)
    else:
        return str(value).strip()


def _create_markdown_table(rows: List[List[str]]) -> str:
    """Create a markdown table from rows of data.

    Args:
        rows: List of rows, where each row is a list of cell values

    Returns:
        Markdown table string
    """
    if not rows:
        return ""

    # Calculate column widths
    num_cols = max(len(row) for row in rows)
    col_widths = [0] * num_cols

    for row in rows:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(cell)))

    # Create markdown table
    lines = []

    for row_idx, row in enumerate(rows):
        # Pad row to match column count
        padded_row = row + [''] * (num_cols - len(row))

        # Format cells with padding
        formatted_cells = [
            str(cell).ljust(col_widths[i])
            for i, cell in enumerate(padded_row)
        ]

        # Create row
        lines.append("| " + " | ".join(formatted_cells) + " |")

        # Add separator after first row (header)
        if row_idx == 0:
            separator = "| " + " | ".join(["-" * width for width in col_widths]) + " |"
            lines.append(separator)

    return "\n".join(lines)
