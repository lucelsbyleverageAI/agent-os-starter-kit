"""Excel processor service for extracting calculated values from Excel files."""

import logging
from typing import List, Optional
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from langchain_core.documents import Document

logger = logging.getLogger(__name__)


class ExcelProcessorService:
    """Service for processing Excel files with calculated value extraction."""

    def process_excel_file(
        self,
        file_path: str,
        include_formulas: bool = False
    ) -> tuple[str, dict]:
        """Process an Excel file and extract content with calculated values.

        Args:
            file_path: Path to the Excel file
            include_formulas: Whether to include formula info in comments (default: False)

        Returns:
            Tuple of (content_markdown, metadata)
        """
        try:
            # Load workbook with data_only=True to get calculated values
            workbook = load_workbook(filename=file_path, data_only=True, read_only=True)

            # Process all sheets
            all_sheets_content = []
            metadata = {
                "file_type": "excel",
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
            }

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_content = self._process_sheet(sheet, sheet_name)

                if sheet_content.strip():  # Only include non-empty sheets
                    all_sheets_content.append(sheet_content)

            # Combine all sheets
            if all_sheets_content:
                content = "\n\n---\n\n".join(all_sheets_content)
            else:
                content = "Empty Excel file"

            workbook.close()

            return content, metadata

        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {e}", exc_info=True)
            raise

    def _process_sheet(self, sheet, sheet_name: str) -> str:
        """Process a single Excel sheet and convert to markdown.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            Markdown representation of the sheet
        """
        lines = []

        # Add sheet header
        lines.append(f"# Sheet: {sheet_name}")
        lines.append("")

        # Find the actual data range (excluding empty rows/columns)
        data_rows = []
        max_col = 0

        for row in sheet.iter_rows(values_only=True):
            # Convert row values, handling None
            row_values = [self._format_cell_value(val) for val in row]

            # Check if row has any non-empty values
            if any(val for val in row_values):
                data_rows.append(row_values)
                # Track maximum column with data
                for i, val in enumerate(row_values):
                    if val:
                        max_col = max(max_col, i + 1)

        if not data_rows:
            lines.append("_Empty sheet_")
            return "\n".join(lines)

        # Trim rows to max_col
        trimmed_rows = [row[:max_col] for row in data_rows]

        # Convert to markdown table
        if trimmed_rows:
            # Create table
            table_md = self._create_markdown_table(trimmed_rows)
            lines.append(table_md)

        return "\n".join(lines)

    def _format_cell_value(self, value) -> str:
        """Format a cell value for display.

        Args:
            value: Cell value (can be None, number, string, etc.)

        Returns:
            Formatted string representation
        """
        if value is None:
            return ""
        elif isinstance(value, (int, float)):
            # Format numbers nicely
            if isinstance(value, float):
                # Check if it's a whole number
                if value.is_integer():
                    return str(int(value))
                else:
                    # Format with appropriate precision
                    return f"{value:.2f}".rstrip('0').rstrip('.')
            return str(value)
        else:
            return str(value).strip()

    def _create_markdown_table(self, rows: List[List[str]]) -> str:
        """Create a markdown table from rows of data.

        Args:
            rows: List of rows, where each row is a list of cell values

        Returns:
            Markdown table string
        """
        if not rows:
            return ""

        # Calculate column widths
        num_cols = max(len(row) for row in rows)
        col_widths = [0] * num_cols

        for row in rows:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(cell)))

        # Create markdown table
        lines = []

        for row_idx, row in enumerate(rows):
            # Pad row to match column count
            padded_row = row + [''] * (num_cols - len(row))

            # Format cells with padding
            formatted_cells = [
                str(cell).ljust(col_widths[i])
                for i, cell in enumerate(padded_row)
            ]

            # Create row
            lines.append("| " + " | ".join(formatted_cells) + " |")

            # Add separator after first row (header)
            if row_idx == 0:
                separator = "| " + " | ".join(["-" * width for width in col_widths]) + " |"
                lines.append(separator)

        return "\n".join(lines)

    async def process_excel_to_documents(
        self,
        file_path: str,
        title: str,
        description: str
    ) -> List[Document]:
        """Process Excel file and return LangChain documents.

        Args:
            file_path: Path to Excel file
            title: Document title
            description: Document description

        Returns:
            List of Document objects (one per sheet or one combined)
        """
        try:
            content, metadata = self.process_excel_file(file_path)

            # Create comprehensive metadata
            doc_metadata = {
                "source": file_path,
                "source_type": "excel",
                "title": title,
                "description": description,
                "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content_length": len(content),
                "word_count": len(content.split()),
                **metadata
            }

            # Create single document with all content
            document = Document(
                page_content=content,
                metadata=doc_metadata
            )

            return [document]

        except Exception as e:
            logger.error(f"Error creating documents from Excel file: {e}", exc_info=True)
            raise


# Global instance
excel_processor_service = ExcelProcessorService()
